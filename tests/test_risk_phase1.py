import importlib
import sys
from copy import deepcopy
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


@pytest.fixture()
def app_module(monkeypatch):
    monkeypatch.setenv("FLASK_SECRET_KEY", "test-secret")
    monkeypatch.setenv("SUPABASE_URL", "https://example.supabase.co")
    monkeypatch.setenv("SUPABASE_ANON_KEY", "anon-test-key")
    monkeypatch.setenv("GEMINI_API_KEY", "gemini-test-key")
    monkeypatch.setenv("FLASK_ENV", "development")

    module = importlib.import_module("riskGenie.app")
    return importlib.reload(module)


@pytest.fixture()
def app(app_module):
    return app_module.create_app({"TESTING": True, "SECRET_KEY": "test-secret"})


@pytest.fixture()
def client(app):
    return app.test_client()


def login_as(client):
    with client.session_transaction() as sess:
        sess["logged_in"] = True
        sess["user_id"] = "user-id"
        sess["username"] = "Peggy"
        sess["email"] = "peggy@example.com"
        sess["role_name"] = "user"
        sess["company_id"] = 7


class FakeQuery:
    def __init__(self, client, table_name):
        self.client = client
        self.table_name = table_name
        self.filters = []
        self.limit_value = None
        self.operation = "select"
        self.insert_payload = None

    def select(self, fields):
        self.operation = "select"
        self.selected_fields = fields
        return self

    def insert(self, payload):
        self.operation = "insert"
        self.insert_payload = deepcopy(payload)
        return self

    def eq(self, field, value):
        self.filters.append(("eq", field, value))
        return self

    def in_(self, field, values):
        self.filters.append(("in", field, list(values)))
        return self

    def limit(self, value):
        self.limit_value = value
        return self

    def order(self, *_args, **_kwargs):
        return self

    def execute(self):
        if self.operation == "insert":
            inserted = deepcopy(self.insert_payload)
            self.client.records.setdefault(self.table_name, []).append(inserted)
            self.client.queries.append({
                "table": self.table_name,
                "filters": list(self.filters),
                "operation": "insert",
                "payload": deepcopy(inserted),
            })
            return SimpleNamespace(data=[inserted])

        records = [
            record.copy()
            for record in self.client.records.get(self.table_name, [])
        ]

        for operation, field, value in self.filters:
            if operation == "eq":
                records = [
                    record
                    for record in records
                    if record.get(field) == value
                ]
            else:
                records = [
                    record
                    for record in records
                    if record.get(field) in value
                ]

        if self.limit_value is not None:
            records = records[:self.limit_value]

        self.client.queries.append({
            "table": self.table_name,
            "filters": list(self.filters),
            "operation": "select",
            "payload": None,
        })
        return SimpleNamespace(data=records)


class FakeSupabase:
    def __init__(self, assets=None, assessments=None):
        self.records = {
            "assets": list(assets or []),
            "risk_assessments": list(assessments or []),
            "audit_logs": [],
        }
        self.queries = []

    def table(self, table_name):
        return FakeQuery(self, table_name)


def install_fake_supabase(monkeypatch, assets=None, assessments=None):
    from riskGenie.models import supabase_db
    from riskGenie.services import risk_routes

    fake = FakeSupabase(assets=assets, assessments=assessments)
    monkeypatch.setattr(risk_routes, "get_supabase_client", lambda: fake)
    monkeypatch.setattr(supabase_db, "get_supabase_client", lambda: fake)
    return fake


def asset_record(asset_id=701, company_id=7, **overrides):
    record = {
        "id": asset_id,
        "company_id": company_id,
        "asset_id_code": f"ASSET-{asset_id}",
        "asset_name": "Web Server",
        "asset_type": "伺服器",
        "description": "Production web server",
        "confidentiality": 5,
        "integrity": 4,
        "availability": 5,
        "legality": 3,
        "asset_value": 5,
        "is_deleted": False,
    }
    record.update(overrides)
    return record


def valid_ai_payload():
    return {
        "asset_id": 701,
        "company_id": 99,
        "asset_name": "Web Server",
        "confidentiality": 5,
        "integrity": 4,
        "availability": 5,
        "legality": 3,
        "cvss_score": 9.8,
        "likelihood_score": 5,
        "impact_score": 9.8,
        "risk_score": 49,
        "risk_level": "極高風險",
    }


def valid_assessment_payload(**overrides):
    payload = {
        "asset_id": 701,
        "company_id": 99,
        "threat_description": "公開服務存在高風險弱點",
        "impact_score": 9.8,
        "likelihood_score": 5,
        "cvss_score": 9.8,
        "risk_score": 49,
        "risk_level": "極高風險",
    }
    payload.update(overrides)
    return payload


def test_weighted_average_and_weighted_avg_match():
    from riskGenie.models.risk_engine import RiskEngine

    weights = {"c": 0.8, "i": 0.1, "a": 0.1}

    weighted_average = RiskEngine.calculate_risk(
        1, 1, 5, 10, "weighted_average", weights
    )
    weighted_avg = RiskEngine.calculate_risk(
        1, 1, 5, 10, "weighted_avg", weights
    )

    assert weighted_average == weighted_avg


def test_weighted_average_does_not_fall_back_to_max():
    from riskGenie.models.risk_engine import RiskEngine

    weights = {"c": 0.8, "i": 0.1, "a": 0.1}

    weighted_average = RiskEngine.calculate_risk(
        1, 1, 5, 10, "weighted_average", weights
    )
    max_value = RiskEngine.calculate_risk(1, 1, 5, 10, "max", weights)

    assert weighted_average == 14
    assert weighted_average != max_value


def test_save_assessment_requires_login(client):
    response = client.post(
        "/api/risk-assessments/save",
        json=valid_assessment_payload(),
    )

    assert response.status_code == 401
    assert response.get_json()["code"] == "UNAUTHORIZED"


def test_save_assessment_requires_company_before_model(client, monkeypatch):
    from riskGenie.services import risk_routes

    monkeypatch.setattr(
        risk_routes,
        "save_risk_assessment_record",
        lambda *_args, **_kwargs: pytest.fail(
            "Risk assessment model must not be called without company context."
        ),
    )
    login_as(client)
    with client.session_transaction() as sess:
        sess.pop("company_id")

    response = client.post(
        "/api/risk-assessments/save",
        json=valid_assessment_payload(),
    )

    assert response.status_code == 403
    assert response.get_json()["code"] == "COMPANY_CONTEXT_REQUIRED"


@pytest.mark.parametrize("asset_id", [True, False, 0, -1, "701", 7.0])
def test_save_assessment_rejects_invalid_asset_id(
    client, monkeypatch, asset_id
):
    from riskGenie.services import risk_routes

    monkeypatch.setattr(
        risk_routes,
        "save_risk_assessment_record",
        lambda *_args, **_kwargs: pytest.fail(
            "Invalid asset_id must be rejected before model access."
        ),
    )
    login_as(client)

    response = client.post(
        "/api/risk-assessments/save",
        json=valid_assessment_payload(asset_id=asset_id),
    )

    assert response.status_code == 400
    assert response.get_json()["code"] == "INVALID_ASSET_ID"


def test_save_assessment_uses_session_company(client, monkeypatch):
    fake = install_fake_supabase(
        monkeypatch,
        assets=[asset_record()],
    )
    login_as(client)

    response = client.post(
        "/api/risk-assessments/save",
        json=valid_assessment_payload(company_id=99),
    )

    assert response.status_code == 201
    saved = fake.records["risk_assessments"]
    assert len(saved) == 1
    assert saved[0]["asset_id"] == 701
    assert saved[0]["company_id"] == 7

    asset_query = next(
        query
        for query in fake.queries
        if query["table"] == "assets"
    )
    assessment_insert = next(
        query
        for query in fake.queries
        if query["table"] == "risk_assessments"
        and query["operation"] == "insert"
    )
    assert ("eq", "id", 701) in asset_query["filters"]
    assert ("eq", "company_id", 7) in asset_query["filters"]
    assert assessment_insert["payload"]["company_id"] == 7


@pytest.mark.parametrize(
    "assets",
    [
        [],
        [asset_record(company_id=99)],
    ],
    ids=["missing-asset", "cross-company-asset"],
)
def test_save_assessment_rejects_unavailable_asset_without_insert(
    client, monkeypatch, assets
):
    fake = install_fake_supabase(monkeypatch, assets=assets)
    login_as(client)

    response = client.post(
        "/api/risk-assessments/save",
        json=valid_assessment_payload(),
    )

    assert response.status_code == 404
    assert response.get_json()["code"] == "ASSET_NOT_FOUND"
    assert fake.records["risk_assessments"] == []
    assert not any(
        query["table"] == "risk_assessments"
        and query["operation"] == "insert"
        for query in fake.queries
    )


def test_save_assessment_db_error_does_not_leak_details(
    client, monkeypatch
):
    from riskGenie.services import risk_routes

    def fail_save(*_args, **_kwargs):
        raise RuntimeError(
            "sensitive SQL https://database.example/internal/server/path"
        )

    monkeypatch.setattr(
        risk_routes,
        "save_risk_assessment_record",
        fail_save,
    )
    login_as(client)

    response = client.post(
        "/api/risk-assessments/save",
        json=valid_assessment_payload(),
    )

    assert response.status_code == 503
    body = response.get_json()
    assert body["code"] == "SAVE_ASSESSMENT_FAILED"
    assert "sensitive" not in str(body)
    assert "https://" not in str(body)


def test_assessment_history_requires_company_before_model(
    client, monkeypatch
):
    from riskGenie.services import risk_routes

    monkeypatch.setattr(
        risk_routes,
        "get_all_risk_assessment_records",
        lambda *_args, **_kwargs: pytest.fail(
            "Risk assessment model must not be called without company context."
        ),
    )
    login_as(client)
    with client.session_transaction() as sess:
        sess.pop("company_id")

    response = client.get("/api/risk-assessments?company_id=99")

    assert response.status_code == 403
    assert response.get_json()["code"] == "COMPANY_CONTEXT_REQUIRED"


def test_assessment_history_is_scoped_to_session_company(
    client, monkeypatch
):
    fake = install_fake_supabase(
        monkeypatch,
        assets=[
            asset_record(),
            asset_record(
                asset_id=9901,
                company_id=99,
                asset_name="Other Company Asset",
            ),
        ],
        assessments=[
            {
                "id": 1,
                "asset_id": 701,
                "company_id": 7,
                "risk_score": 49,
                "created_at": "2026-08-21T00:00:00+00:00",
            },
            {
                "id": 2,
                "asset_id": 9901,
                "company_id": 99,
                "risk_score": 99,
                "created_at": "2026-08-22T00:00:00+00:00",
            },
            {
                "id": 3,
                "asset_id": 9901,
                "company_id": 7,
                "risk_score": 88,
                "created_at": "2026-08-23T00:00:00+00:00",
            },
        ],
    )
    login_as(client)

    response = client.get("/api/risk-assessments?company_id=99")

    assert response.status_code == 200
    assessments = response.get_json()["assessments"]
    assert [assessment["id"] for assessment in assessments] == [1]
    assert "Other Company Asset" not in str(assessments)

    assessment_query = next(
        query
        for query in fake.queries
        if query["table"] == "risk_assessments"
        and query["operation"] == "select"
    )
    asset_query = next(
        query
        for query in fake.queries
        if query["table"] == "assets"
    )
    assert ("eq", "company_id", 7) in assessment_query["filters"]
    assert ("eq", "company_id", 7) in asset_query["filters"]
    assert ("in", "id", [701, 9901]) in asset_query["filters"]


def test_assessment_history_db_error_does_not_leak_details(
    client, monkeypatch
):
    from riskGenie.services import risk_routes

    def fail_history(*_args, **_kwargs):
        raise RuntimeError(
            "sensitive SQL https://database.example/internal/server/path"
        )

    monkeypatch.setattr(
        risk_routes,
        "get_all_risk_assessment_records",
        fail_history,
    )
    login_as(client)

    response = client.get("/api/risk-assessments")

    assert response.status_code == 503
    body = response.get_json()
    assert body["code"] == "FETCH_ASSESSMENTS_FAILED"
    assert "sensitive" not in str(body)
    assert "https://" not in str(body)


def test_ai_advice_requires_login(client):
    response = client.post("/api/ai-advice", json=valid_ai_payload())

    assert response.status_code == 401
    assert response.is_json
    assert response.get_json()["code"] == "UNAUTHORIZED"


def test_ai_advice_page_scopes_assets_and_assessments_to_session_company(
    client, monkeypatch
):
    from riskGenie.services import risk_routes

    fake = install_fake_supabase(
        monkeypatch,
        assets=[
            asset_record(),
            asset_record(asset_id=9901, company_id=99),
        ],
        assessments=[
            {
                "id": 1,
                "asset_id": 701,
                "company_id": 7,
            },
            {
                "id": 2,
                "asset_id": 9901,
                "company_id": 99,
            },
        ],
    )
    rendered = {}
    monkeypatch.setattr(
        risk_routes,
        "render_template",
        lambda template, **context: (
            rendered.update(template=template, **context) or "rendered"
        ),
    )
    login_as(client)

    response = client.get("/ai-advice?company_id=999")

    assert response.status_code == 200
    assert rendered == {
        "template": "ai_advice.html",
        "has_assessment": True,
    }
    asset_query, assessment_query = fake.queries
    assert ("eq", "company_id", 7) in asset_query["filters"]
    assert ("eq", "is_deleted", False) in asset_query["filters"]
    assert ("eq", "company_id", 7) in assessment_query["filters"]
    assert ("in", "asset_id", [701]) in assessment_query["filters"]


def test_ai_advice_page_ignores_deleted_assets_and_skips_assessment_query(
    client, monkeypatch
):
    from riskGenie.services import risk_routes

    fake = install_fake_supabase(
        monkeypatch,
        assets=[asset_record(is_deleted=True)],
        assessments=[
            {
                "id": 1,
                "asset_id": 701,
                "company_id": 7,
            },
        ],
    )
    rendered = {}
    monkeypatch.setattr(
        risk_routes,
        "render_template",
        lambda template, **context: (
            rendered.update(template=template, **context) or "rendered"
        ),
    )
    login_as(client)

    response = client.get("/ai-advice")

    assert response.status_code == 200
    assert rendered["has_assessment"] is False
    assert len(fake.queries) == 1
    assert fake.queries[0]["table"] == "assets"
    assert ("eq", "company_id", 7) in fake.queries[0]["filters"]
    assert ("eq", "is_deleted", False) in fake.queries[0]["filters"]


def test_ai_advice_page_ignores_cross_company_and_mismatched_assessments(
    client, monkeypatch
):
    from riskGenie.services import risk_routes

    fake = install_fake_supabase(
        monkeypatch,
        assets=[asset_record()],
        assessments=[
            {
                "id": 1,
                "asset_id": 701,
                "company_id": 99,
            },
            {
                "id": 2,
                "asset_id": 9901,
                "company_id": 7,
            },
        ],
    )
    rendered = {}
    monkeypatch.setattr(
        risk_routes,
        "render_template",
        lambda template, **context: (
            rendered.update(template=template, **context) or "rendered"
        ),
    )
    login_as(client)

    response = client.get("/ai-advice")

    assert response.status_code == 200
    assert rendered["has_assessment"] is False
    assessment_query = fake.queries[1]
    assert assessment_query["table"] == "risk_assessments"
    assert ("eq", "company_id", 7) in assessment_query["filters"]
    assert ("in", "asset_id", [701]) in assessment_query["filters"]


def test_export_requires_login(client, monkeypatch):
    from riskGenie.services import risk_routes

    monkeypatch.setattr(
        risk_routes,
        "get_supabase_client",
        lambda: pytest.fail("Supabase must not be queried."),
    )

    response = client.get("/export")

    assert response.status_code == 401
    assert response.is_json
    assert response.get_json()["code"] == "UNAUTHORIZED"


def test_export_requires_company_before_query(client, monkeypatch):
    from riskGenie.services import risk_routes

    monkeypatch.setattr(
        risk_routes,
        "get_supabase_client",
        lambda: pytest.fail("Supabase must not be queried."),
    )
    login_as(client)
    with client.session_transaction() as sess:
        sess.pop("company_id")

    response = client.get("/export")

    assert response.status_code == 403
    assert response.get_json()["code"] == "COMPANY_CONTEXT_REQUIRED"


def test_ai_advice_rejects_non_json(client):
    login_as(client)

    response = client.post(
        "/api/ai-advice",
        data="not-json",
        content_type="text/plain",
    )

    assert response.status_code == 400
    assert response.get_json()["code"] == "INVALID_JSON"


def test_ai_advice_requires_asset_id(client):
    login_as(client)

    response = client.post("/api/ai-advice", json={})

    assert response.status_code == 400
    assert response.get_json()["code"] == "MISSING_ASSET_ID"


@pytest.mark.parametrize("asset_id", [True, 0, -1, "invalid"])
def test_ai_advice_rejects_invalid_asset_id(client, asset_id):
    login_as(client)

    response = client.post(
        "/api/ai-advice",
        json={"asset_id": asset_id},
    )

    assert response.status_code == 400
    assert response.get_json()["code"] == "INVALID_ASSET_ID"


def test_ai_advice_requires_completed_risk_assessment(
    client, monkeypatch
):
    from riskGenie.services import risk_routes

    install_fake_supabase(monkeypatch, assets=[asset_record()])
    monkeypatch.setattr(
        risk_routes,
        "is_gemini_configured",
        lambda: pytest.fail("Gemini config must not be checked."),
    )
    monkeypatch.setattr(
        risk_routes,
        "generate_advice",
        lambda **_kwargs: pytest.fail("RAG/Gemini must not be called."),
    )
    login_as(client)
    payload = valid_ai_payload()
    del payload["risk_level"]

    response = client.post("/api/ai-advice", json=payload)

    assert response.status_code == 409
    assert response.get_json()["code"] == "RISK_ASSESSMENT_REQUIRED"


def test_ai_advice_rejects_cross_company_asset(client, monkeypatch):
    from riskGenie.services import risk_routes

    fake = install_fake_supabase(
        monkeypatch,
        assets=[asset_record(company_id=99)],
    )
    monkeypatch.setattr(
        risk_routes,
        "generate_advice",
        lambda **_kwargs: pytest.fail("RAG/Gemini must not be called."),
    )
    login_as(client)

    response = client.post(
        "/api/ai-advice",
        json=valid_ai_payload(),
    )

    assert response.status_code == 404
    assert response.get_json()["code"] == "ASSET_NOT_FOUND"
    assert ("eq", "company_id", 7) in fake.queries[0]["filters"]


def test_ai_advice_without_gemini_key_fails_safely(
    client, monkeypatch
):
    from riskGenie.services import risk_routes

    def fail_if_called(**_kwargs):
        raise AssertionError("Gemini must not be called without GEMINI_API_KEY.")

    install_fake_supabase(monkeypatch, assets=[asset_record()])
    monkeypatch.delenv("GEMINI_API_KEY", raising=False)
    monkeypatch.setattr(risk_routes, "generate_advice", fail_if_called)
    login_as(client)

    response = client.post("/api/ai-advice", json=valid_ai_payload())

    assert response.status_code == 503
    body = response.get_json()
    assert body["success"] is False
    assert body["code"] == "GEMINI_NOT_CONFIGURED"
    assert body["error"] == "AI 服務尚未設定"


@pytest.mark.parametrize("api_key", ["", "   "])
def test_blank_gemini_key_is_not_configured(monkeypatch, api_key):
    from riskGenie.services import rag_service

    monkeypatch.setenv("GEMINI_API_KEY", api_key)

    assert rag_service.is_gemini_configured() is False


def test_ai_advice_exception_does_not_leak_details(client, monkeypatch):
    from riskGenie.services import risk_routes

    install_fake_supabase(monkeypatch, assets=[asset_record()])
    monkeypatch.setattr(risk_routes, "is_gemini_configured", lambda: True)

    def fail_generation(**_kwargs):
        raise RuntimeError(
            "sensitive SDK detail https://internal.example/server/path"
        )

    monkeypatch.setattr(risk_routes, "generate_advice", fail_generation)
    login_as(client)

    response = client.post("/api/ai-advice", json=valid_ai_payload())

    assert response.status_code == 503
    body = response.get_json()
    assert body == {
        "success": False,
        "error": "AI 建議產生失敗，請稍後再試。",
        "code": "AI_GENERATION_FAILED",
    }
    assert "sensitive" not in str(body)
    assert "https://" not in str(body)


def test_ai_advice_uses_session_company_and_returns_advice(
    client, monkeypatch
):
    from riskGenie.services import risk_routes

    fake = install_fake_supabase(monkeypatch, assets=[asset_record()])
    monkeypatch.setattr(risk_routes, "is_gemini_configured", lambda: True)
    monkeypatch.setattr(
        risk_routes,
        "generate_advice",
        lambda **_kwargs: "請優先修補公開服務。",
    )
    login_as(client)

    response = client.post("/api/ai-advice", json=valid_ai_payload())

    assert response.status_code == 200
    assert response.get_json()["advice"] == "請優先修補公開服務。"
    assert ("eq", "company_id", 7) in fake.queries[0]["filters"]
    assert ("eq", "is_deleted", False) in fake.queries[0]["filters"]


def test_export_only_contains_current_company_data(client, monkeypatch):
    fake = install_fake_supabase(
        monkeypatch,
        assets=[
            asset_record(),
            asset_record(
                asset_id=702,
                asset_name="Deleted Asset",
                is_deleted=True,
            ),
            asset_record(
                asset_id=9901,
                company_id=99,
                asset_name="Other Company Asset",
            ),
        ],
        assessments=[
            {
                "id": 1,
                "asset_id": 701,
                "company_id": 7,
                "cvss_score": 9.8,
                "likelihood_score": 5,
                "impact_score": 9.8,
                "risk_score": 49,
                "risk_level": "極高風險",
                "created_at": "2026-08-21T00:00:00+00:00",
            },
            {
                "id": 2,
                "asset_id": 9901,
                "company_id": 99,
                "risk_score": 99,
                "risk_level": "Other Company",
            },
            {
                "id": 3,
                "asset_id": 9901,
                "company_id": 7,
                "risk_score": 88,
                "risk_level": "Mismatched Asset",
            },
            {
                "id": 4,
                "asset_id": 702,
                "company_id": 7,
                "risk_score": 77,
                "risk_level": "Deleted Asset",
            },
        ],
    )
    login_as(client)

    response = client.get("/export")

    assert response.status_code == 200
    assert response.mimetype == (
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.data), read_only=True)
    rows = list(workbook["風險報表"].iter_rows(values_only=True))
    workbook.close()

    assert rows[0][:3] == ("資產代碼", "資產名稱", "資產類型")
    assert len(rows) == 2
    assert rows[1][1] == "Web Server"
    serialized_rows = str(rows)
    assert "Other Company Asset" not in serialized_rows
    assert "Deleted Asset" not in serialized_rows
    assert "Mismatched Asset" not in serialized_rows

    asset_query, assessment_query = fake.queries
    assert ("eq", "company_id", 7) in asset_query["filters"]
    assert ("eq", "is_deleted", False) in asset_query["filters"]
    assert ("eq", "company_id", 7) in assessment_query["filters"]
    assert ("in", "asset_id", [701]) in assessment_query["filters"]


def test_weight_settings_reports_local_backup_when_supabase_fails(
    tmp_path, monkeypatch
):
    from riskGenie.services import risk_service

    def fail_supabase():
        raise RuntimeError("database connection details must stay server-side")

    monkeypatch.setattr(risk_service, "get_supabase_client", fail_supabase)
    monkeypatch.setattr(
        risk_service,
        "FALLBACK_FILE",
        str(tmp_path / "weight_settings_fallback.json"),
    )

    result = risk_service.RiskService.save_weight_settings(
        company_id=7,
        formula_type="weighted_average",
        weight_c=0.4,
        weight_i=0.3,
        weight_a=0.3,
    )

    assert result["success"] is True
    assert result["supabase_synced"] is False
    assert result["local_backup_saved"] is True
    assert result["status"] == "local_backup_only"
